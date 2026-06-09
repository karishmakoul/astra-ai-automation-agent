"""
Excel Reader: reads our AmbitionBox_Test_Cases.xlsx (or any sheet with the same schema).
No LLM call needed — the data is already structured.

Expected columns (by position, 1-based):
  1: TC ID       2: Title      3: Priority   4: Type
  5: Preconditions  6: Test Steps  7: Expected Result  8: Automation Status
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font

from ai_agent.models import TestSpec, TestCase, Priority, TestType, InputSource

# ── Status cell styles ────────────────────────────────────────────────────────
_GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_WHITE_FILL  = PatternFill(fill_type=None)   # reset to no fill

_STATUS_STYLES = {
    "Automated":          (_GREEN_FILL,  Font(color="276221")),
    "Automated (Failing)":(_YELLOW_FILL, Font(color="9C5700")),
    "To Be Automated":    (_WHITE_FILL,  Font()),
}


# Maps column header names to our Priority enum
_PRIORITY_MAP = {
    "critical": Priority.CRITICAL,
    "high":     Priority.HIGH,
    "medium":   Priority.MEDIUM,
    "low":      Priority.LOW,
}

_TYPE_MAP = {
    "smoke":      TestType.SMOKE,
    "regression": TestType.REGRESSION,
    "sanity":     TestType.SANITY,
}


def _cell(row, col: int) -> str:
    """Safe cell read — returns empty string if cell is None."""
    val = row[col - 1].value
    return str(val).strip() if val is not None else ""


def _parse_steps(raw: str) -> list[str]:
    """Split numbered steps or newline-separated steps into a list."""
    if not raw:
        return []
    lines = [l.strip() for l in raw.splitlines() if l.strip()]
    return lines


def read_excel(
    file_path: str,
    sheet_name: str | None = None,
    status_filter: str | None = "To Be Automated",
) -> TestSpec:
    """
    Read test cases from the Excel file and return a TestSpec.

    Args:
        file_path:     Path to the .xlsx file
        sheet_name:    Specific sheet to read (None = read all sheets combined)
        status_filter: Only include rows with this Automation Status value.
                       Pass None to include all rows.

    Returns:
        TestSpec populated from the Excel rows
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Decide which sheets to read
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        sheets_to_read = [sheet_name]
    else:
        sheets_to_read = wb.sheetnames

    all_test_cases: list[TestCase] = []
    affected_pages: list[str] = []

    for sname in sheets_to_read:
        ws = wb[sname]
        # Derive a snake_case page name from the sheet name
        page_name = sname.lower().replace(" ", "_")
        rows = list(ws.iter_rows())

        if len(rows) < 2:
            continue  # skip empty sheets

        # Skip header row (row 0)
        for row in rows[1:]:
            tc_id    = _cell(row, 1)
            title    = _cell(row, 2)
            priority = _cell(row, 3)
            tc_type  = _cell(row, 4)
            precon   = _cell(row, 5)
            steps    = _cell(row, 6)
            expected = _cell(row, 7)
            status   = _cell(row, 8)

            if not tc_id or not title:
                continue  # blank row

            if status_filter and status.lower() != status_filter.lower():
                continue  # filtered out

            all_test_cases.append(
                TestCase(
                    id=tc_id,
                    title=title,
                    priority=_PRIORITY_MAP.get(priority.lower(), Priority.HIGH),
                    type=_TYPE_MAP.get(tc_type.lower(), TestType.REGRESSION),
                    preconditions=_parse_steps(precon),
                    steps=_parse_steps(steps),
                    expected_result=expected,
                    tags=[page_name],
                )
            )

        if any(tc for tc in all_test_cases if page_name in tc.tags):
            affected_pages.append(page_name)

    wb.close()   # read_only workbook

    sheet_label = sheet_name or "all sheets"
    filter_label = f" (status='{status_filter}')" if status_filter else ""

    return TestSpec(
        source=InputSource.EXCEL,
        source_id=str(path),
        title=f"Test cases from {path.name} — {sheet_label}{filter_label}",
        description=(
            f"Loaded {len(all_test_cases)} test cases from "
            f"'{path.name}' ({sheet_label}){filter_label}"
        ),
        affected_pages=list(dict.fromkeys(affected_pages)),  # dedup, preserve order
        user_types=[],
        test_cases=all_test_cases,
        raw_content=f"Excel file: {path}",
    )


def update_status_in_excel(
    file_path: str,
    sheet_name: str | None,
    tc_ids: list[str],
    new_status: str = "Automated",
) -> int:
    """
    Update the Automation Status column (col 8) for matching TC IDs.

    Opens the workbook in write mode, finds every row whose TC ID (col 1)
    is in `tc_ids`, sets the status cell to `new_status`, applies the
    matching colour style, and saves.

    Args:
        file_path:  Path to the .xlsx file
        sheet_name: Sheet(s) to update. None = all sheets.
        tc_ids:     List of TC ID strings to mark (e.g. ["SAL_001", "SAL_002"])
        new_status: The status text to write (default "Automated")

    Returns:
        Number of cells updated
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    tc_id_set = set(tc_ids)
    fill, font = _STATUS_STYLES.get(new_status, (_GREEN_FILL, Font(color="276221")))

    # openpyxl write mode (NOT read_only)
    wb = openpyxl.load_workbook(path, data_only=True)

    sheets_to_update = [sheet_name] if sheet_name else wb.sheetnames
    updated = 0

    for sname in sheets_to_update:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        rows = list(ws.iter_rows(min_row=2))   # skip header
        for row in rows:
            tc_id_cell = row[0]                # col A = TC ID
            status_cell = row[7]               # col H = Automation Status

            if tc_id_cell.value and str(tc_id_cell.value).strip() in tc_id_set:
                status_cell.value = new_status
                status_cell.fill  = fill
                status_cell.font  = font
                updated += 1

    wb.save(path)
    wb.close()
    return updated
